import openpyxl
import json
from datetime import datetime

# Leer el Excel
excel_path = r"C:\Users\Daniel\Downloads\productos almacen de herrajes 5-6.xlsx"
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# Obtener encabezados
headers = []
for cell in ws[1]:
    headers.append(cell.value)

print("=" * 80)
print("ANÁLISIS DEL EXCEL DE PRODUCTOS")
print("=" * 80)
print("\nColumnas encontradas:")
for i, h in enumerate(headers):
    print(f"  {i}: {h}")

# Mostrar primeras 10 filas
print("\nPrimeros 10 productos:")
print("-" * 80)
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=11, values_only=True), 1):
    print(f"Fila {row_idx}:")
    for col_idx, (header, value) in enumerate(zip(headers, row)):
        if value is not None:
            print(f"  {header}: {value}")

print(f"\nTotal de filas de datos: {ws.max_row - 1}")
print("\n" + "=" * 80)
